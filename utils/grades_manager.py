# -*- coding: utf-8 -*-
import sqlite3
import os
import pandas as pd
from datetime import datetime
import traceback

class GradesManager:
    def __init__(self, db_path="students.db"):
        self.db_path = db_path
        self._ensure_table_exists()
    
    def _ensure_table_exists(self):
        """确保学生表存在并包含必要的成绩字段"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 检查学生表结构
        cursor.execute("PRAGMA table_info(students)")
        columns = cursor.fetchall()
        column_names = [column[1] for column in columns]
        
        # 需要的成绩字段
        grade_fields = [
            'daof', 'yuwen', 'shuxue', 'yingyu', 'laodong', 
            'tiyu', 'yinyue', 'meishu', 'kexue', 'zonghe', 
            'xinxi', 'shufa'
        ]
        
        # 添加缺失的字段
        for field in grade_fields:
            if field not in column_names:
                cursor.execute(f"ALTER TABLE students ADD COLUMN {field} TEXT DEFAULT ''")
        
        # 确保存在学期字段
        if 'semester' not in column_names:
            cursor.execute("ALTER TABLE students ADD COLUMN semester TEXT DEFAULT '上学期'")
        
        conn.commit()
        conn.close()
    
    def get_all_grades(self, semester="上学期"):
        """获取所有学生的成绩"""
        print(f"获取所有学生成绩，学期: {semester}")
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            # 直接从students表获取全部数据，不对学期做过滤
            cursor.execute('''
                SELECT 
                    id AS student_id, 
                    name AS student_name, 
                    class, 
                    daof, yuwen, shuxue, yingyu, laodong, 
                    tiyu, yinyue, meishu, kexue, zonghe, 
                    xinxi, shufa
                FROM students
                ORDER BY class, CAST(id AS INTEGER)
            ''')
            
            # 获取列名
            column_names = [description[0] for description in cursor.description]
            
            # 获取数据
            all_students = [dict(zip(column_names, row)) for row in cursor.fetchall()]
            
            # 为每个学生记录添加学期字段
            for student in all_students:
                student['semester'] = semester
            
            print(f"找到 {len(all_students)} 个学生记录")
            return all_students
            
        except Exception as e:
            print(f"获取所有学生成绩时出错: {e}")
            print(traceback.format_exc())
            return []
        finally:
            conn.close()
    
    def get_student_grade(self, student_id, semester="上学期"):
        """获取单个学生的成绩记录"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # 查询学生信息和成绩
        cursor.execute('''
            SELECT id, name, class, 
                   daof, yuwen, shuxue, yingyu, laodong, 
                   tiyu, yinyue, meishu, kexue, zonghe, 
                   xinxi, shufa, 
                   COALESCE(semester, ?) AS semester
            FROM students 
            WHERE id = ? AND (semester = ? OR semester IS NULL OR semester = '')
        ''', (semester, student_id, semester))
        
        row = cursor.fetchone()
        conn.close()
        
        if row:
            # 构造字典
            return {
                'student_id': row[0],
                'student_name': row[1],
                'class': row[2],
                'daof': row[3] or '',
                'yuwen': row[4] or '',
                'shuxue': row[5] or '',
                'yingyu': row[6] or '',
                'laodong': row[7] or '',
                'tiyu': row[8] or '',
                'yinyue': row[9] or '',
                'meishu': row[10] or '',
                'kexue': row[11] or '',
                'zonghe': row[12] or '',
                'xinxi': row[13] or '',
                'shufa': row[14] or '',
                'semester': row[15]
            }
        else:
            # 查找学生基本信息
            cursor.execute('SELECT id, name, class FROM students WHERE id = ?', (student_id,))
            student = cursor.fetchone()
            
            if student:
                return {
                    'student_id': student[0],
                    'student_name': student[1],
                    'class': student[2],
                    'semester': semester,
                    'daof': '',
                    'yuwen': '',
                    'shuxue': '',
                    'yingyu': '',
                    'laodong': '',
                    'tiyu': '',
                    'yinyue': '',
                    'meishu': '',
                    'kexue': '',
                    'zonghe': '',
                    'xinxi': '',
                    'shufa': ''
                }
            return None  # 学生不存在
    
    def save_grade(self, student_id, grade_data, semester="上学期"):
        """保存单个学生的成绩记录"""
        print(f"准备保存学生 {student_id} 的成绩，学期: {semester}")
        print(f"成绩数据: {grade_data}")
        
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        try:
            # 检查学生是否存在
            cursor.execute("SELECT id FROM students WHERE id = ?", (student_id,))
            if not cursor.fetchone():
                print(f"学生ID {student_id} 不存在")
                return False
                
            # 生成更新字句
            set_clauses = []
            values = []
            
            for field in ['daof', 'yuwen', 'shuxue', 'yingyu', 'laodong', 
                        'tiyu', 'yinyue', 'meishu', 'kexue', 'zonghe', 
                        'xinxi', 'shufa']:
                if field in grade_data:
                    value = grade_data.get(field, '')
                    
                    # 验证成绩值是否符合「优、良、及格、待及格」四级制
                    if value and value not in ['优', '良', '及格', '待及格', '差', '']:
                        print(f"警告: {student_id} 的 {field} 成绩 '{value}' 不符合要求，已自动清空")
                        value = ''  # 清空不符合要求的值
                        
                    set_clauses.append(f"{field} = ?")
                    values.append(value)
            
            # 添加更新时间
            set_clauses.append("updated_at = ?")
            values.append(now)
            
            # 添加条件参数
            values.append(student_id)
            
            # 执行更新
            if set_clauses:
                query = f'''
                UPDATE students SET {', '.join(set_clauses)}
                WHERE id = ?
                '''
                print(f"执行SQL: {query}")
                print(f"参数: {values}")
                cursor.execute(query, values)
                
                print(f"更新了 {cursor.rowcount} 条记录")
                conn.commit()
                print(f"成功保存学生 {student_id} 的成绩")
                return True
            else:
                print("没有提供任何成绩数据")
                return False
            
        except Exception as e:
            print(f"保存成绩时出错: {e}")
            print(traceback.format_exc())
            conn.rollback()
            return False
        finally:
            conn.close()
    
    def delete_grade(self, student_id, semester="上学期"):
        """清空学生的成绩记录"""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        try:
            # 清空成绩字段
            cursor.execute('''
            UPDATE students SET
                daof = '',
                yuwen = '',
                shuxue = '',
                yingyu = '',
                laodong = '',
                tiyu = '',
                yinyue = '',
                meishu = '',
                kexue = '',
                zonghe = '',
                xinxi = '',
                shufa = '',
                semester = ?
            WHERE id = ? AND (semester = ? OR semester IS NULL OR semester = '')
            ''', (semester, student_id, semester))
            
            conn.commit()
            return True
        except Exception as e:
            print(f"删除成绩时出错: {e}")
            conn.rollback()
            return False
        finally:
            conn.close()
    
    def import_grades_from_excel(self, file_path, semester="上学期"):
        """从Excel文件导入成绩"""
        try:
            # 检查文件路径
            if not file_path or not os.path.exists(file_path):
                print(f"文件路径无效或文件不存在: {file_path}")
                return False, f"文件路径无效或文件不存在: {file_path}"
                
            print(f"准备导入成绩文件: {file_path}, 文件大小: {os.path.getsize(file_path)} 字节")
            
            # 读取Excel文件
            df = pd.read_excel(file_path)
            print(f"成功读取Excel文件，包含 {len(df)} 行数据")
            
            # 确保有必要的列
            required_columns = ['学号']
            for col in required_columns:
                if col not in df.columns:
                    return False, f"Excel文件缺少必要的列: {col}"
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            # 获取列名映射
            column_mapping = self._get_column_mapping()
            
            # 允许的成绩值
            allowed_grades = ['优', '良', '及格', '待及格', '差', '']
            
            # 成绩值标准化映射表
            grade_mapping = {
                '优秀': '优',
                '良好': '良',
                '及': '及格',
                '待': '待及格'
            }
            
            # 检查是否包含姓名列，用于校验
            has_name_column = '姓名' in df.columns
            name_mismatch_count = 0
            
            # 先检查Excel中的列是否能够识别
            recognized_subjects = []
            unrecognized_columns = []
            
            for col in df.columns:
                if col in ['学号', '姓名']:
                    continue
                if col in column_mapping:
                    recognized_subjects.append(col)
                else:
                    unrecognized_columns.append(col)
            
            if unrecognized_columns:
                print(f"发现未识别的列: {unrecognized_columns}，这些列将被忽略")
            
            if not recognized_subjects:
                return False, f"Excel文件中没有可识别的成绩列。可用的科目包括：{', '.join(column_mapping.keys())}"
            
            # 转换Excel数据
            success_count = 0
            fail_count = 0
            skipped_count = 0
            warnings = []
            
            for i, row in df.iterrows():
                try:
                    student_id = str(row['学号']).strip()
                    
                    # 检查学生是否存在
                    cursor.execute("SELECT id, name FROM students WHERE id = ?", (student_id,))
                    student_record = cursor.fetchone()
                    if not student_record:
                        print(f"学生ID不存在: {student_id}")
                        fail_count += 1
                        continue
                    
                    db_student_name = student_record[1]
                    
                    # 如果Excel有姓名列，则检查姓名是否匹配
                    if has_name_column:
                        excel_student_name = str(row['姓名']).strip() if pd.notna(row['姓名']) else ""
                        
                        if excel_student_name and excel_student_name != db_student_name:
                            name_mismatch_msg = f"姓名不匹配: 学号 {student_id} 在Excel中为「{excel_student_name}」，系统中为「{db_student_name}」"
                            print(name_mismatch_msg)
                            warnings.append(name_mismatch_msg)
                            name_mismatch_count += 1
                            fail_count += 1
                            continue
                    
                    # 生成更新字句
                    set_clauses = ["semester = ?"]
                    values = [semester]
                    
                    # 记录是否有有效成绩数据
                    has_valid_grade = False
                    
                    for excel_col in recognized_subjects:
                        db_col = column_mapping[excel_col]
                        value = str(row[excel_col]) if pd.notna(row[excel_col]) else ''
                        
                        # 验证成绩值是否在允许的范围内
                        if value:
                            # 尝试标准化成绩值
                            if value in grade_mapping:
                                value = grade_mapping[value]
                            
                            if value not in allowed_grades:
                                print(f"警告: 学号 {student_id} 的 {excel_col} 成绩 '{value}' 不符合要求，已自动跳过")
                                value = ''  # 跳过不符合要求的成绩
                            else:
                                has_valid_grade = True  # 找到至少一个有效成绩
                        
                        set_clauses.append(f"{db_col} = ?")
                        values.append(value)
                    
                    # 只有有至少一个有效成绩才更新数据库
                    if has_valid_grade:
                        # 添加条件参数
                        values.append(student_id)
                        
                        # 执行更新
                        if len(set_clauses) > 1:  # 确保至少有一个成绩字段要更新
                            query = f'''
                            UPDATE students SET {', '.join(set_clauses)}
                            WHERE id = ?
                            '''
                            cursor.execute(query, values)
                            
                            if cursor.rowcount > 0:
                                success_count += 1
                            else:
                                fail_count += 1
                    else:
                        # 没有有效成绩，跳过此记录
                        print(f"学号 {student_id} 没有有效的成绩数据，已跳过")
                        skipped_count += 1
                except Exception as e:
                    print(f"导入学生 {student_id} 的成绩时出错: {e}")
                    fail_count += 1
            
            conn.commit()
            conn.close()
            
            print(f"导入完成: 成功 {success_count} 条, 跳过 {skipped_count} 条, 失败 {fail_count} 条")
            
            # 删除临时文件
            try:
                if os.path.exists(file_path):
                    os.remove(file_path)
                    print(f"已删除临时文件: {file_path}")
            except Exception as e:
                print(f"删除临时文件失败: {e}")
            
            status_message = f"成功导入 {success_count} 条成绩记录"
            if skipped_count > 0:
                status_message += f"，跳过 {skipped_count} 条无效记录"
            if fail_count > 0:
                status_message += f"，失败 {fail_count} 条"
            if name_mismatch_count > 0:
                status_message += f"，其中 {name_mismatch_count} 条因姓名不匹配而跳过"
            
            return True, status_message
        
        except Exception as e:
            print(f"导入成绩时出错: {e}")
            print(traceback.format_exc())  # 打印完整的错误堆栈
            return False, f"导入成绩时出错: {str(e)}"
    
    def preview_grades_from_excel(self, file_path, semester="上学期"):
        """从Excel文件预览成绩导入，不实际导入数据库"""
        try:
            # 检查文件路径
            if not file_path or not os.path.exists(file_path):
                print(f"文件路径无效或文件不存在: {file_path}")
                return {
                    'status': 'error',
                    'message': f"文件路径无效或文件不存在: {file_path}"
                }
                
            print(f"准备预览成绩文件: {file_path}, 文件大小: {os.path.getsize(file_path)} 字节")
            
            # 读取Excel文件
            df = pd.read_excel(file_path)
            print(f"成功读取Excel文件，包含 {len(df)} 行数据")
            
            # 确保有必要的列
            required_columns = ['学号']
            for col in required_columns:
                if col not in df.columns:
                    return {
                        'status': 'error',
                        'message': f"Excel文件缺少必要的列: {col}"
                    }
            
            # 获取列名映射
            column_mapping = self._get_column_mapping()
            
            # 允许的成绩值
            allowed_grades = ['优', '良', '及格', '待及格', '差', '']
            
            # 成绩值标准化映射表
            grade_mapping = {
                '优秀': '优',
                '良好': '良',
                '及': '及格',
                '待': '待及格'
            }
            
            # 从数据库获取学生信息
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('SELECT id, name, class FROM students')
            students_dict = {row[0]: {'name': row[1], 'class': row[2]} for row in cursor.fetchall()}
            conn.close()
            
            # 转换Excel数据为预览数据
            preview_data = []
            valid_count = 0
            invalid_count = 0
            warnings = []
            name_mismatches = []  # 记录姓名不匹配的情况
            recognized_subjects = []  # 记录成功识别的科目
            unrecognized_columns = []  # 记录未识别的列
            
            # 检查是否包含姓名列，用于校验
            has_name_column = '姓名' in df.columns
            
            # 先检查Excel中的列是否能够识别
            for col in df.columns:
                if col in ['学号', '姓名']:
                    continue
                if col in column_mapping:
                    recognized_subjects.append(col)
                else:
                    unrecognized_columns.append(col)
            
            if unrecognized_columns:
                print(f"发现未识别的列: {unrecognized_columns}")
            
            if not recognized_subjects:
                return {
                    'status': 'error',
                    'message': f"Excel文件中没有可识别的成绩列。可用的科目包括：{', '.join(column_mapping.keys())}"
                }
            
            for i, row in df.iterrows():
                try:
                    student_id = str(row['学号']).strip()
                    
                    # 检查学生是否存在
                    if student_id not in students_dict:
                        warnings.append(f"警告: 学号 {student_id} 不在系统中")
                        invalid_count += 1
                        continue
                    
                    # 如果Excel有姓名列，则检查姓名是否匹配
                    excel_student_name = ""
                    if has_name_column:
                        excel_student_name = str(row['姓名']).strip() if pd.notna(row['姓名']) else ""
                        db_student_name = students_dict[student_id]['name']
                        
                        if excel_student_name and excel_student_name != db_student_name:
                            name_mismatch_msg = f"姓名不匹配: 学号 {student_id} 在Excel中为「{excel_student_name}」，系统中为「{db_student_name}」"
                            warnings.append(name_mismatch_msg)
                            name_mismatches.append({
                                'student_id': student_id,
                                'excel_name': excel_student_name,
                                'db_name': db_student_name
                            })
                            # 姓名不匹配视为无效记录
                            invalid_count += 1
                            continue
                    
                    # 创建学生成绩记录
                    student_grade = {
                        'student_id': student_id,
                        'student_name': students_dict[student_id]['name'],
                        'class': students_dict[student_id]['class'],
                        'semester': semester
                    }
                    
                    # 添加各科目成绩
                    has_valid_grade = False  # 是否有至少一个有效成绩
                    grade_warnings = []  # 该学生的成绩警告
                    
                    for excel_col in recognized_subjects:
                        db_col = column_mapping[excel_col]
                        value = str(row[excel_col]) if pd.notna(row[excel_col]) else ''
                        
                        # 验证成绩值是否在允许的范围内
                        if value:
                            # 尝试标准化成绩值
                            if value in grade_mapping:
                                value = grade_mapping[value]
                            
                            if value not in allowed_grades:
                                grade_warnings.append(f"学号 {student_id} 的 {excel_col} 成绩 '{value}' 不符合要求，已自动跳过")
                                value = ''  # 在预览中标记为空
                            else:
                                has_valid_grade = True  # 找到至少一个有效成绩
                        
                        student_grade[db_col] = value
                    
                    # 只有有效的成绩才添加到预览数据中
                    if has_valid_grade:
                        preview_data.append(student_grade)
                        valid_count += 1
                        
                        # 添加该学生的警告（如果有）
                        warnings.extend(grade_warnings)
                    else:
                        # 此学生没有有效成绩，计入无效记录
                        warnings.append(f"学号 {student_id} 没有有效的成绩数据，已跳过")
                        invalid_count += 1
                
                except Exception as e:
                    warnings.append(f"处理学号 {student_id} 的成绩时出错: {str(e)}")
                    invalid_count += 1
            
            # 生成HTML预览
            html_preview = self._generate_grades_preview_html(preview_data, warnings, recognized_subjects, unrecognized_columns, name_mismatches)
            
            status_message = f'成功识别 {len(recognized_subjects)} 个科目的成绩，共 {valid_count} 条有效记录'
            if unrecognized_columns:
                status_message += f'，跳过了 {len(unrecognized_columns)} 个无法识别的列'
            if invalid_count > 0:
                status_message += f'，{invalid_count} 条记录无效'
            if name_mismatches:
                status_message += f'，发现 {len(name_mismatches)} 条姓名不匹配'
            
            return {
                'status': 'ok',
                'message': status_message,
                'grades': preview_data,
                'html_preview': html_preview,
                'warnings': warnings,
                'file_path': file_path,
                'recognized_subjects': recognized_subjects,
                'unrecognized_columns': unrecognized_columns,
                'name_mismatches': name_mismatches
            }
            
        except Exception as e:
            print(f"预览成绩时出错: {e}")
            print(traceback.format_exc())
            return {
                'status': 'error',
                'message': f"预览成绩时出错: {str(e)}"
            }
    
    def _generate_grades_preview_html(self, grades_data, warnings=None, recognized_subjects=None, unrecognized_columns=None, name_mismatches=None):
        """生成成绩预览的HTML表格"""
        print("生成成绩预览HTML表格")
        
        # 定义科目名称字典
        subject_names = self.get_subject_names()
        
        # 添加识别状态信息
        html = ""
        if recognized_subjects or unrecognized_columns:
            html += """
            <div class="alert alert-info mb-3">
                <h5><i class='bx bx-info-circle'></i> 导入信息：</h5>
            """
            
            if recognized_subjects:
                html += f"""
                <p><strong>已识别的科目({len(recognized_subjects)})：</strong></p>
                <div style="display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 12px;">
                """
                
                # 使用绿色荧光背景显示已识别的科目
                for subject in recognized_subjects:
                    html += f"""
                    <span style="display: inline-block; padding: 4px 10px; background-color: #a7f3d0; color: #047857; border-radius: 4px; font-weight: 500;">{subject}</span>
                    """
                
                html += "</div>"
            
            if unrecognized_columns:
                html += f"""
                <p><strong>未识别的列({len(unrecognized_columns)})：</strong> <small class="text-muted">(这些列将被忽略)</small></p>
                <div style="display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 12px;">
                """
                
                # 使用红色荧光背景显示未识别的列
                for column in unrecognized_columns:
                    html += f"""
                    <span style="display: inline-block; padding: 4px 10px; background-color: #fecaca; color: #b91c1c; border-radius: 4px; font-weight: 500;">{column}</span>
                    """
                
                html += "</div>"
                
            html += "</div>"
        
        # 添加姓名不匹配警告
        if name_mismatches and len(name_mismatches) > 0:
            html += """
            <div class="alert alert-danger mb-3">
                <h5><i class='bx bx-error-circle'></i> 姓名不匹配警告：</h5>
                <p class="mb-2">以下学生的姓名与系统中的记录不一致，这些学生的成绩将被<strong>跳过导入</strong>。请检查姓名是否正确。</p>
                
                <div class="table-responsive mt-2">
                    <table class="table table-sm table-bordered table-danger">
                        <thead>
                            <tr>
                                <th>学号</th>
                                <th>Excel中的姓名</th>
                                <th>系统中的姓名</th>
                            </tr>
                        </thead>
                        <tbody>
            """
            
            for mismatch in name_mismatches:
                html += f"""
                <tr>
                    <td>{mismatch['student_id']}</td>
                    <td style="background-color: #fee2e2;">{mismatch['excel_name']}</td>
                    <td>{mismatch['db_name']}</td>
                </tr>
                """
            
            html += """
                        </tbody>
                    </table>
                </div>
                <p class="mt-2 mb-0"><strong>解决方法：</strong>请检查Excel文件中的姓名拼写，确保与系统中的学生姓名完全一致。</p>
            </div>
            """
        
        html += """
        <div class="table-responsive">
            <table class="table table-striped table-hover">
                <thead>
                    <tr>
                        <th>学号</th>
                        <th>姓名</th>
                        <th>班级</th>
        """
        
        # 添加科目列标题 - 仅包含已识别的科目
        if recognized_subjects:
            for subject in ['道法', '语文', '数学', '英语', '劳动', '体育', '音乐', 
                            '美术', '科学', '综合', '信息', '书法']:
                if subject in recognized_subjects:
                    html += f"""
                        <th style="background-color: #d1fae5;">{subject}</th>
                    """
        else:
            # 如果没有指定已识别的科目，显示所有科目列
            html += """
                        <th>道法</th>
                        <th>语文</th>
                        <th>数学</th>
                        <th>英语</th>
                        <th>劳动</th>
                        <th>体育</th>
                        <th>音乐</th>
                        <th>美术</th>
                        <th>科学</th>
                        <th>综合</th>
                        <th>信息</th>
                        <th>书法</th>
            """
            
        html += """
                    </tr>
                </thead>
                <tbody>
        """
        
        # 添加成绩数据行
        for grade in grades_data:
            html += f"""
                <tr>
                    <td>{grade.get('student_id', '-')}</td>
                    <td>{grade.get('student_name', '-')}</td>
                    <td>{grade.get('class', '-')}</td>
            """
            
            # 添加各科目成绩 - 仅包含已识别的科目
            if recognized_subjects:
                for subject_code, subject_name in subject_names.items():
                    if subject_name in recognized_subjects:
                        html += f"""
                    <td>{grade.get(subject_code, '-')}</td>
                        """
            else:
                # 如果没有指定已识别的科目，显示所有科目成绩
                html += f"""
                    <td>{grade.get('daof', '-')}</td>
                    <td>{grade.get('yuwen', '-')}</td>
                    <td>{grade.get('shuxue', '-')}</td>
                    <td>{grade.get('yingyu', '-')}</td>
                    <td>{grade.get('laodong', '-')}</td>
                    <td>{grade.get('tiyu', '-')}</td>
                    <td>{grade.get('yinyue', '-')}</td>
                    <td>{grade.get('meishu', '-')}</td>
                    <td>{grade.get('kexue', '-')}</td>
                    <td>{grade.get('zonghe', '-')}</td>
                    <td>{grade.get('xinxi', '-')}</td>
                    <td>{grade.get('shufa', '-')}</td>
                """
                
            html += """
                </tr>
            """
        
        html += """
                </tbody>
            </table>
        </div>
        """
        
        # 添加警告信息（如果有）
        if warnings and len(warnings) > 0:
            # 限制显示的警告数量，避免太多警告淹没界面
            max_warnings = 10
            warning_count = len(warnings)
            
            html += """
            <div class="alert alert-warning mt-3">
                <h5><i class='bx bx-error'></i> 警告信息：</h5>
            """
            
            if warning_count <= max_warnings:
                html += "<ul>"
                for warning in warnings:
                    html += f"<li>{warning}</li>"
                html += "</ul>"
            else:
                # 只显示部分警告，并提示总数
                html += f"<p>共有 {warning_count} 条警告，显示前 {max_warnings} 条：</p><ul>"
                for warning in warnings[:max_warnings]:
                    html += f"<li>{warning}</li>"
                html += f"</ul><p>...还有 {warning_count - max_warnings} 条警告未显示</p>"
            
            html += """
            </div>
            """
        
        # 添加数据统计
        html += f"""
        <div class="alert alert-success mt-3">
            <i class='bx bx-check-circle'></i> 共发现 {len(grades_data)} 条有效成绩数据，点击"确认导入"按钮完成导入。
            <p class="mb-0 mt-2"><small>系统已自动识别可导入的科目，并跳过了无法识别的列和无效的成绩数据。</small></p>
        </div>
        """
        
        return html
    
    def _get_column_mapping(self):
        """获取成绩Excel列名到数据库字段的映射"""
        return {
            '学号': 'student_id',
            '道法': 'daof',
            '语文': 'yuwen',
            '数学': 'shuxue',
            '英语': 'yingyu',
            '劳动': 'laodong',
            '体育': 'tiyu',
            '音乐': 'yinyue',
            '美术': 'meishu',
            '科学': 'kexue',
            '综合': 'zonghe',
            '信息': 'xinxi',
            '书法': 'shufa'
        }
    
    def get_subject_names(self):
        """获取所有科目名称"""
        return {
            'daof': '道法',
            'yuwen': '语文',
            'shuxue': '数学',
            'yingyu': '英语',
            'laodong': '劳动',
            'tiyu': '体育',
            'yinyue': '音乐',
            'meishu': '美术',
            'kexue': '科学',
            'zonghe': '综合',
            'xinxi': '信息',
            'shufa': '书法'
        }
        
    def create_empty_template(self, output_path=None):
        """创建空白的成绩导入模板"""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 获取所有学生
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
        SELECT id, name, class FROM students
        ORDER BY class, CAST(id AS INTEGER)
        ''')
        
        students = []
        for row in cursor.fetchall():
            students.append({
                'id': row[0],
                'name': row[1],
                'class': row[2]
            })
        
        conn.close()
        
        # 创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "成绩导入模板"
        
        # 设置列标题
        columns = [
            {'header': '学号', 'key': 'student_id', 'width': 15},
            {'header': '姓名', 'key': 'student_name', 'width': 15},
            {'header': '班级', 'key': 'class', 'width': 15},
            {'header': '道法', 'key': 'daof', 'width': 10},
            {'header': '语文', 'key': 'yuwen', 'width': 10},
            {'header': '数学', 'key': 'shuxue', 'width': 10},
            {'header': '英语', 'key': 'yingyu', 'width': 10},
            {'header': '劳动', 'key': 'laodong', 'width': 10},
            {'header': '体育', 'key': 'tiyu', 'width': 10},
            {'header': '音乐', 'key': 'yinyue', 'width': 10},
            {'header': '美术', 'key': 'meishu', 'width': 10},
            {'header': '科学', 'key': 'kexue', 'width': 10},
            {'header': '综合', 'key': 'zonghe', 'width': 10},
            {'header': '信息', 'key': 'xinxi', 'width': 10},
            {'header': '书法', 'key': 'shufa', 'width': 10},
        ]
        
        # 设置标题行
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=column['header'])
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.column_dimensions[get_column_letter(col_idx)].width = column['width']
        
        # 添加学生信息
        for row_idx, student in enumerate(students, start=2):
            ws.cell(row=row_idx, column=1, value=student['id'])
            ws.cell(row=row_idx, column=2, value=student['name'])
            ws.cell(row=row_idx, column=3, value=student['class'])
        
        # 添加说明文字
        notes_row = len(students) + 3
        ws.cell(row=notes_row, column=1, value="说明事项：")
        ws.cell(row=notes_row+1, column=1, value="1. 成绩可填写：优、良、及格、待及格，请勿使用其他评分方式")
        ws.cell(row=notes_row+2, column=1, value="2. 请勿修改学号、姓名和班级")
        ws.cell(row=notes_row+3, column=1, value="3. 导入时只需填写有变更的成绩")
        
        # 合并说明文字的单元格
        for i in range(4):
            ws.merge_cells(start_row=notes_row+i, start_column=1, end_row=notes_row+i, end_column=6)
        
        # 确定保存路径
        if not output_path:
            # 使用默认路径
            template_dir = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'templates')
            os.makedirs(template_dir, exist_ok=True)
            output_path = os.path.join(template_dir, 'grades_import_template.xlsx')
        
        try:
            # 保存文件
            wb.save(output_path)
            print(f"成绩导入模板已保存到: {output_path}")
            return output_path
        except Exception as e:
            print(f"保存成绩导入模板时出错: {e}")
            print(traceback.format_exc())
            
            # 尝试保存到备用位置
            backup_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'grades_template.xlsx')
            try:
                wb.save(backup_path)
                print(f"成绩导入模板已保存到备用位置: {backup_path}")
                return backup_path
            except Exception as e2:
                print(f"保存到备用位置也失败: {e2}")
                return None

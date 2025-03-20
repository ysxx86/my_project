# -*- coding: utf-8 -*-
# 自动安装所需依赖
import sys
import subprocess
import os
import json

def check_and_install(package, version=None):
    """检查并安装Python包"""
    package_with_version = f"{package}=={version}" if version else package
    
    try:
        # 尝试导入模块
        __import__(package)
        print(f"✓ {package} 已安装")
    except ImportError:
        print(f"! 未找到 {package} 模块，正在安装...")
        try:
            # 使用系统Python运行pip安装
            subprocess.check_call([sys.executable, "-m", "pip", "install", 
                                  package_with_version, "--no-cache-dir"])
            print(f"✓ {package} 安装成功")
        except Exception as e:
            print(f"! {package} 安装失败: {str(e)}")
            print(f"  请手动运行: pip install {package_with_version}")

# 检查关键依赖
print("检查并安装关键依赖...")
REQUIRED_PACKAGES = [
    ("flask", "3.0.2"),
    ("flask_cors", "3.0.10"),
    ("pandas", "2.2.1"),
    ("openpyxl", "3.1.2"),
    ("werkzeug", "2.3.0")
]

# 单独处理requests和reportlab
try:
    import requests
    print("✓ requests 已安装")
except ImportError:
    print("! 未找到 requests 模块，评语AI生成功能将不可用")
    print("  如需使用AI生成评语，请运行: pip install requests==2.31.0")

try:
    import reportlab
    print("✓ reportlab 已安装")
except ImportError:
    print("! 未找到 reportlab 模块，PDF导出功能将不可用")
    print("  如需使用PDF导出功能，请运行: pip install reportlab==4.1.0")

# 安装基础依赖
for package, version in REQUIRED_PACKAGES:
    check_and_install(package, version)

print("依赖检查完成，开始导入模块...\n")

# 原始的导入语句
from flask import Flask, request, jsonify, send_from_directory, render_template, url_for, send_file
from flask_cors import CORS
import os
import sqlite3
import pandas as pd
import datetime
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import shutil
import time
import logging
import traceback
from utils.excel_processor import ExcelProcessor
from utils.comment_processor import batch_update_comments, generate_comments_pdf, generate_preview_html
try:
    from utils.pdf_exporter import export_comments_to_pdf  # 导入PDF导出函数
except ImportError:
    # 创建一个简化版的PDF导出函数
    def export_comments_to_pdf(*args, **kwargs):
        return None, "PDF导出功能不可用，请安装reportlab模块"
    print("! PDF导出功能不可用，请安装reportlab模块")
    
from utils.grades_manager import GradesManager
from utils.comment_generator import CommentGenerator

# 设置DeepSeek API全局变量
DEEPSEEK_API_KEY = os.environ.get("DEEPSEEK_API_KEY", "")
deepseek_api = None

# 尝试导入和初始化DeepSeek API
try:
    from utils.deepseek_api import DeepSeekAPI
    if DEEPSEEK_API_KEY:
        deepseek_api = DeepSeekAPI(DEEPSEEK_API_KEY)
        print("✓ DeepSeek API 已初始化")
    else:
        deepseek_api = DeepSeekAPI(None)
        print("! DeepSeek API密钥未设置，AI评语生成功能不可用")
except ImportError:
    print("! 无法导入DeepSeekAPI，AI评语生成功能不可用")
    deepseek_api = None

# 系统设置存储
SYSTEM_SETTINGS = {
    "deepseek_api_key": DEEPSEEK_API_KEY,
    "deepseek_api_enabled": bool(DEEPSEEK_API_KEY)
}

# 配置日志
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler("server.log"),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)
logger.info("服务器启动")

# 创建Flask应用，指定静态文件夹
app = Flask(__name__, 
            static_url_path='', 
            static_folder='./',
            template_folder='./')
CORS(app)  # 启用跨域资源共享

# 配置
UPLOAD_FOLDER = 'uploads'
TEMPLATE_FOLDER = 'templates'
EXPORTS_FOLDER = 'exports'
DATABASE = 'students.db'

# 将UPLOAD_FOLDER添加到app.config中
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# 确保所有必要的文件夹存在
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(TEMPLATE_FOLDER, exist_ok=True)
os.makedirs(EXPORTS_FOLDER, exist_ok=True)

# 创建数据库连接
def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

# 初始化数据库
def init_db():
    logger.info("初始化数据库")
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    
    # 创建学生表 - 移除了birth_date、phone、address和notes字段
    cursor.execute('''
    CREATE TABLE IF NOT EXISTS students (
        id TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        gender TEXT NOT NULL,
        class TEXT,
        height REAL,
        weight REAL,
        chest_circumference REAL,
        vital_capacity REAL,
        dental_caries TEXT,
        vision_left REAL,
        vision_right REAL,
        physical_test_status TEXT,
        comments TEXT,
        created_at TEXT,
        updated_at TEXT
    )
    ''')
    
    # 检查是否所有必要的列都存在，如果不存在则添加
    cursor.execute("PRAGMA table_info(students)")
    existing_columns = [row[1] for row in cursor.fetchall()]
    logger.info(f"现有数据库列: {existing_columns}")
    
    # 定义应该存在的列及其类型
    expected_columns = {
        'chest_circumference': 'REAL',
        'vital_capacity': 'REAL',
        'dental_caries': 'TEXT',
        'vision_left': 'REAL',
        'vision_right': 'REAL',
        'physical_test_status': 'TEXT',
        'comments': 'TEXT'
    }
    
    # 添加缺失的列
    for column, col_type in expected_columns.items():
        if column not in existing_columns:
            logger.warning(f"添加缺失的列: {column} ({col_type})")
            try:
                cursor.execute(f"ALTER TABLE students ADD COLUMN {column} {col_type}")
                conn.commit()
                logger.info(f"成功添加列: {column}")
            except sqlite3.Error as e:
                logger.error(f"添加列 {column} 时出错: {e}")
    
    # 提交并关闭连接
    conn.commit()
    conn.close()
    
    logger.info("数据库初始化完成")

# 重置数据库功能
def reset_db():
    """重置数据库，备份旧数据库并创建新的"""
    try:
        # 如果数据库存在，创建备份
        if os.path.exists(DATABASE):
            backup_path = f"{DATABASE}.backup_{int(time.time())}"
            shutil.copy2(DATABASE, backup_path)
            logger.info(f"创建数据库备份: {backup_path}")
            
            # 删除旧数据库
            os.remove(DATABASE)
            logger.info(f"删除旧数据库: {DATABASE}")
        
        # 初始化新数据库
        init_db()
        return True, "数据库已成功重置，旧数据已备份"
    except Exception as e:
        logger.error(f"重置数据库时出错: {e}")
        logger.error(traceback.format_exc())
        return False, f"重置数据库失败: {str(e)}"

# 初始化数据库
init_db()

# 创建学生导入模板
def create_student_template():
    template_dir = 'templates'
    if not os.path.exists(template_dir):
        os.makedirs(template_dir)
    
    template_path = os.path.join(template_dir, 'student_template.xlsx')
    
    # 如果文件存在且被占用，则跳过创建
    if os.path.exists(template_path):
        try:
            # 尝试打开文件，如果可以打开就先删除
            with open(template_path, 'a'):
                pass
            os.remove(template_path)
        except:
            print("学生模板文件被占用，跳过创建")
            return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "学生信息"
    
    # 设置标题行
    headers = ['学号', '姓名', '性别', '班级', '身高(cm)', '体重(kg)', '胸围(cm)', '肺活量(ml)', '龋齿(个)', '左眼视力', '右眼视力']
    for i, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(i)].width = 15
    
    # 添加示例数据
    example_data = ['1', '张三', '男', '三年级一班', '135', '32', '65', '1500', '0', '5.0', '5.0']
    for i, value in enumerate(example_data, 1):
        ws.cell(row=2, column=i, value=value)
    
    # 添加说明文字
    ws.cell(row=4, column=1, value="说明事项：")
    ws.cell(row=5, column=1, value="1. 请按照示例格式填写学生信息")
    ws.cell(row=6, column=1, value='2. 性别请填写"男"或"女"')
    ws.cell(row=7, column=1, value="3. 班级格式: 三年级一班")
    ws.cell(row=8, column=1, value="4. 视力格式: 5.0 或 4.8 等")
    
    # 合并说明文字的单元格
    for i in range(4, 9):
        ws.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
    
    wb.save(template_path)
    print("学生Excel模板创建完成")

# 创建模板文件
try:
    create_student_template()
except Exception as e:
    print(f"创建学生模板时出错: {str(e)}")

# 创建成绩管理器实例
grades_manager = GradesManager()

# 创建成绩导入模板
try:
    grades_manager.create_empty_template()
    print("成绩Excel模板创建完成")
except Exception as e:
    print(f"创建成绩模板时出错: {str(e)}")

# 全局变量
DATABASE_FILE = os.environ.get('CLASS_MASTER_DB', 'classmaster.db')
LOG_FILE = os.environ.get('CLASS_MASTER_LOG', 'classmaster.log')

# 主页路由
@app.route('/')
def index():
    return send_from_directory('./', 'index.html')

# 页面路由
@app.route('/pages/<path:path>')
def serve_pages(path):
    return send_from_directory('pages', path)

# 健康检查API
@app.route('/api/health', methods=['GET'])
def health_check():
    return jsonify({'status': 'ok', 'message': '服务器正常运行中'})

# 下载模板API
@app.route('/api/template', endpoint='download_student_template', methods=['GET'])
def download_template():
    template_path = os.path.join(TEMPLATE_FOLDER, 'student_template.xlsx')
    if not os.path.exists(template_path):
        create_student_template()
    
    return jsonify({
        'status': 'ok',
        'message': '模板文件准备就绪',
        'template_url': f'/download/template/student_template.xlsx'
    })

# 提供模板下载
@app.route('/download/template/<filename>', methods=['GET'])
def serve_template(filename):
    return send_from_directory(TEMPLATE_FOLDER, filename)

# 导入学生预览API
@app.route('/api/import-students', methods=['POST'])
def import_students_preview():
    if 'file' not in request.files:
        return jsonify({'error': '没有上传文件'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': '没有选择文件'}), 400
    
    if not (file.filename.endswith('.xlsx') or file.filename.endswith('.xls')):
        return jsonify({'error': '只支持.xlsx或.xls格式的Excel文件'}), 400
    
    # 保存上传的文件
    filename = secure_filename(file.filename)
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    saved_filename = f"{timestamp}_{filename}"
    file_path = os.path.join(UPLOAD_FOLDER, saved_filename)
    file.save(file_path)
    
    try:
        # 使用新的Excel处理器处理文件
        processor = ExcelProcessor()
        result = processor.process_file(file_path)
        
        # 处理结果
        if 'error' in result:
            return jsonify({'error': result['error']}), 400
        
        return jsonify({
            'status': 'ok',
            'message': result['message'],
            'html_preview': result['html_preview'],
            'students': result['students'],
            'file_path': file_path
        })
    
    except Exception as e:
        logger.error(f"解析Excel文件时出错: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'error': f'解析Excel文件时出错: {str(e)}'}), 500

# 确认导入学生API
@app.route('/api/confirm-import', methods=['POST'])
def confirm_import():
    data = request.json
    
    # 增加详细日志记录，帮助诊断问题
    logger.info("接收到的导入确认请求数据: %s", data)
    
    if not data or 'students' not in data:
        return jsonify({'error': '无效的请求数据'}), 400
    
    students = data['students']
    
    if not students or len(students) == 0:
        return jsonify({'error': '没有学生数据可导入'}), 400
    
    # 记录第一条学生数据的结构，帮助诊断
    if students and len(students) > 0:
        logger.info("第一条学生数据样例: %s", students[0])
        logger.info("学生数据字段: %s", list(students[0].keys()))
        
        # 详细记录第一条学生的数值字段
        first_student = students[0]
        for field in ['height', 'weight', 'chest_circumference', 'vital_capacity', 'vision_left', 'vision_right']:
            if field in first_student:
                value = first_student[field]
                logger.info(f"数值字段 {field}: {value}, 类型: {type(value).__name__}")
    
    # 检查数据库表结构
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 获取表结构
    cursor.execute("PRAGMA table_info(students)")
    db_columns = [row[1] for row in cursor.fetchall()]
    logger.info("数据库表字段: %s", db_columns)
    
    # 字段名映射表，用于处理代码和数据库字段名不一致的情况
    field_mappings = {
        'chest_circumference': 'chest_circumference',  # 可能的映射
        'chestCircumference': 'chest_circumference',   # 前端JS中可能使用的驼峰命名
        'vital_capacity': 'vital_capacity',
        'vitalCapacity': 'vital_capacity',
        'dental_caries': 'dental_caries',
        'dentalCaries': 'dental_caries',
        'vision_left': 'vision_left',
        'visionLeft': 'vision_left',
        'vision_right': 'vision_right',
        'visionRight': 'vision_right',
        'physical_test_status': 'physical_test_status',
        'physicalTestStatus': 'physical_test_status'
    }
    
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 清空学生表中的所有记录 - 根据用户需求，导入时只保留当前班级的学生
    try:
        logger.info("导入前清空学生表中的所有记录")
        cursor.execute('DELETE FROM students')
        logger.info(f"已清空学生表，准备导入 {len(students)} 名新学生")
    except Exception as clear_error:
        logger.error(f"清空学生表时出错: {str(clear_error)}")
        return jsonify({
            'status': 'error',
            'message': f'清空数据库时出错: {str(clear_error)}'
        }), 500
    
    success_count = 0
    error_count = 0
    updated_count = 0
    inserted_count = 0
    error_details = []
    
    try:
        for i, student in enumerate(students):
            try:
                # 打印学生数据以便调试
                logger.info(f"准备导入的学生数据({i+1}/{len(students)}): {student}")
                
                # 添加数据验证
                if 'id' not in student or not student['id']:
                    raise ValueError(f"学生ID缺失或无效: {student}")
                if 'name' not in student or not student['name']:
                    raise ValueError(f"学生姓名缺失或无效: {student}")
                if 'gender' not in student:
                    raise ValueError(f"学生性别缺失: {student}")
                
                # 准备SQL参数，确保使用正确的字段名
                params = {
                    'id': student['id'],
                    'name': student['name'],
                    'gender': student['gender'],
                    'class': student.get('class', ''),
                    'created_at': now,
                    'updated_at': now
                }
                
                # 数值字段的特殊处理
                numeric_fields = [
                    'height', 'weight', 'chest_circumference', 
                    'vital_capacity', 'vision_left', 'vision_right'
                ]
                
                for field in numeric_fields:
                    # 获取字段值，对null、空字符串、null字符串等进行处理
                    raw_value = student.get(field)
                    if raw_value is None or raw_value == '' or raw_value == 'null' or raw_value == 'undefined':
                        params[field] = None
                    else:
                        # 尝试转换为浮点数
                        try:
                            # 如果是字符串，处理可能的特殊格式
                            if isinstance(raw_value, str):
                                # 替换逗号为点号(小数点)
                                raw_value = raw_value.replace(',', '.')
                            
                            # 转换为浮点数
                            value = float(raw_value)
                            # 特别处理0值
                            params[field] = 0.0 if value == 0 else value
                            
                            logger.info(f"字段 {field} 原始值: {raw_value} ({type(raw_value).__name__}) -> 转换值: {params[field]} ({type(params[field]).__name__})")
                        except (ValueError, TypeError) as e:
                            params[field] = None
                            logger.warning(f"无法转换字段 {field} 的值 {raw_value}: {str(e)}")
                
                # 处理文本字段
                if 'dental_caries' in student:
                    params['dental_caries'] = student['dental_caries']
                if 'physical_test_status' in student:
                    params['physical_test_status'] = student['physical_test_status']
                
                # 处理特殊字段，确保使用正确的数据库字段名 (处理可能的命名不一致)
                for field in student:
                    if field in field_mappings and field_mappings[field] in db_columns:
                        db_field = field_mappings[field]
                        params[db_field] = student[field]
                
                # 打印最终的SQL参数
                logger.info(f"最终SQL参数: {params}")
                
                # 检查学生是否已存在
                cursor.execute('SELECT id FROM students WHERE id = ?', (student['id'],))
                existing_student = cursor.fetchone()
                
                if existing_student:
                    # 构建更新SQL - 修改为覆盖更新所有字段
                    update_fields = []
                    update_values = []
                    
                    # 遍历所有数据库列，不仅仅是导入数据中的列
                    for field in db_columns:
                        if field != 'id':  # ID不更新
                            # 如果字段在导入数据参数中，使用新值
                            # 否则设置为NULL或空字符串（根据字段类型）
                            if field in params:
                                value = params[field]
                            else:
                                # 为文本字段设空字符串，数值字段设NULL
                                is_text_field = field in ['name', 'gender', 'class', 'dental_caries', 
                                                         'physical_test_status', 'comments', 'created_at', 'updated_at']
                                value = '' if is_text_field else None
                                
                                # 特殊处理评语字段，确保清空
                                if field == 'comments':
                                    value = ''
                                    
                            update_fields.append(f"{field} = ?")
                            update_values.append(value)
                    
                    # 添加WHERE条件的参数
                    update_values.append(student['id'])
                    
                    # 执行更新
                    update_sql = f"UPDATE students SET {', '.join(update_fields)} WHERE id = ?"
                    logger.info(f"更新SQL: {update_sql}")
                    logger.info(f"更新参数: {update_values}")
                    cursor.execute(update_sql, update_values)
                    updated_count += 1
                else:
                    # 构建插入SQL - 修改为包含所有字段
                    insert_fields = []
                    insert_values = []
                    insert_placeholders = []
                    
                    # 遍历所有数据库列，不仅仅是导入数据中的列
                    for field in db_columns:
                        insert_fields.append(field)
                        
                        # 如果字段在导入数据参数中，使用新值
                        # 否则设置为NULL或空字符串（根据字段类型）
                        if field in params:
                            value = params[field]
                        else:
                            # 为文本字段设空字符串，数值字段设NULL
                            is_text_field = field in ['name', 'gender', 'class', 'dental_caries', 
                                                     'physical_test_status', 'comments', 'created_at', 'updated_at']
                            value = '' if is_text_field else None
                            
                            # 特殊处理评语字段，确保为空
                            if field == 'comments':
                                value = ''
                        
                        insert_values.append(value)
                        insert_placeholders.append('?')
                    
                    # 执行插入
                    insert_sql = f"INSERT INTO students ({', '.join(insert_fields)}) VALUES ({', '.join(insert_placeholders)})"
                    logger.info(f"插入SQL: {insert_sql}")
                    logger.info(f"插入参数: {insert_values}")
                    cursor.execute(insert_sql, insert_values)
                    inserted_count += 1
                
                success_count += 1
            except Exception as student_error:
                # 单个学生导入错误不应该影响整体事务
                error_count += 1
                error_msg = f"导入学生 {student.get('id', '未知ID')} 时出错: {str(student_error)}"
                logger.error(error_msg)
                error_details.append(error_msg)
                # 继续处理下一个学生，而不是中断整个批次
        
        # 如果至少有一个学生成功导入，提交事务
        if success_count > 0:
            conn.commit()
            logger.info(f"成功导入 {success_count} 名学生，更新 {updated_count} 名，新增 {inserted_count} 名")
        else:
            conn.rollback()
            logger.warning("没有学生成功导入，回滚事务")
    except Exception as e:
        conn.rollback()
        error_count += 1
        error_msg = f"导入学生数据时出错: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        error_details.append(error_msg)
        return jsonify({
            'status': 'error',
            'message': error_msg,
            'error_details': error_details
        }), 500
    finally:
        conn.close()
    
    # 如果有错误但也有成功，返回部分成功的响应
    status = 'ok' if error_count == 0 else 'partial'
    return jsonify({
        'status': status,
        'message': f'成功导入{success_count}名学生（新增{inserted_count}名，更新{updated_count}名）' +
                  (f'，{error_count}名学生导入失败' if error_count > 0 else ''),
        'success_count': success_count,
        'error_count': error_count,
        'updated_count': updated_count,
        'inserted_count': inserted_count,
        'error_details': error_details if error_count > 0 else []
    })

# 获取所有学生API
@app.route('/api/students', methods=['GET'], strict_slashes=False)
def get_all_students():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM students ORDER BY CAST(id AS INTEGER)')
    students = [dict(row) for row in cursor.fetchall()]
    
    conn.close()
    
    return jsonify({
        'status': 'ok',
        'count': len(students),
        'students': students
    })

# 获取单个学生API
@app.route('/api/students/<student_id>', methods=['GET'], strict_slashes=False)
def get_student(student_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('SELECT * FROM students WHERE id = ?', (student_id,))
    student = cursor.fetchone()
    
    conn.close()
    
    if student:
        return jsonify({
            'status': 'ok',
            'student': dict(student)
        })
    else:
        return jsonify({'error': '未找到学生'}), 404

# 添加新学生API
@app.route('/api/students', methods=['POST'], strict_slashes=False)
def add_student():
    data = request.json
    
    if not data:
        return jsonify({'error': '无效的请求数据'}), 400
    
    required_fields = ['id', 'name', 'gender']
    for field in required_fields:
        if field not in data:
            return jsonify({'error': f'缺少必要的字段: {field}'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 检查学生ID是否已存在
    cursor.execute('SELECT id FROM students WHERE id = ?', (data['id'],))
    if cursor.fetchone():
        conn.close()
        return jsonify({'error': f'学号 {data["id"]} 已存在'}), 400
    
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    # 确保数值类型字段的正确处理
    height = data.get('height')
    weight = data.get('weight')
    chest_circumference = data.get('chest_circumference')
    vital_capacity = data.get('vital_capacity')
    vision_left = data.get('vision_left')
    vision_right = data.get('vision_right')
    
    try:
        cursor.execute('''
        INSERT INTO students (
            id, name, gender, class, height, weight,
            chest_circumference, vital_capacity, dental_caries,
            vision_left, vision_right, physical_test_status, created_at, updated_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            data['id'], data['name'], data['gender'], data.get('class', ''),
            height, weight, 
            chest_circumference, vital_capacity, data.get('dental_caries', ''),
            vision_left, vision_right, data.get('physical_test_status', ''),
            now, now
        ))
        
        conn.commit()
        
        return jsonify({
            'status': 'ok',
            'message': f'成功添加学生: {data["name"]}',
            'student_id': data['id']
        })
    except Exception as e:
        conn.rollback()
        print(f"添加学生时出错: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': f'添加学生时出错: {str(e)}'}), 500
    finally:
        conn.close()

# 更新学生API
@app.route('/api/students/<student_id>', methods=['PUT'], strict_slashes=False)
def update_student(student_id):
    logger.info(f"收到更新学生信息请求，学生ID: {student_id}")
    
    try:
        data = request.json
        logger.info(f"请求数据: {data}")
        
        if not data:
            logger.error("无效的请求数据")
            return jsonify({'error': '无效的请求数据'}), 400
        
        # 确保请求的ID与URL中的ID一致
        if 'id' in data and data['id'] != student_id:
            logger.error(f"URL中的ID({student_id})与请求体中的ID({data['id']})不一致")
            return jsonify({'error': '学生ID不一致'}), 400
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 检查学生是否存在
        cursor.execute('SELECT id FROM students WHERE id = ?', (student_id,))
        if not cursor.fetchone():
            conn.close()
            return jsonify({'error': f'未找到学号为 {student_id} 的学生'}), 404
        
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 更详细的数值处理和日志
        numeric_fields = {
            'height': 'height',
            'weight': 'weight',
            'chest_circumference': 'chest_circumference',
            'vital_capacity': 'vital_capacity',
            'vision_left': 'vision_left',
            'vision_right': 'vision_right',
        }
        
        processed_values = {}
        
        # 处理数值字段
        for field, db_field in numeric_fields.items():
            raw_value = data.get(field, None)
            logger.info(f"字段 {field} 原始值: {raw_value}, 类型: {type(raw_value).__name__}")
            
            try:
                if raw_value is None or raw_value == '' or raw_value == 'null' or raw_value == 'undefined':
                    processed_values[db_field] = None
                    logger.info(f"字段 {field} 设为None")
                else:
                    # 如果是字符串，处理逗号等
                    if isinstance(raw_value, str):
                        # 替换逗号为点
                        raw_value = raw_value.replace(',', '.')
                        logger.info(f"字段 {field} 预处理后: {raw_value}")
                    
                    # 转换为浮点数
                    value = float(raw_value)
                    # 处理0值
                    processed_values[db_field] = 0.0 if value == 0 else value
                    logger.info(f"字段 {field} 成功转换为: {processed_values[db_field]}")
            except (ValueError, TypeError) as e:
                logger.warning(f"字段 {field} 值 '{raw_value}' 转换错误: {str(e)}")
                processed_values[db_field] = None
        
        # 检查表结构
        cursor.execute("PRAGMA table_info(students)")
        existing_columns = [row[1] for row in cursor.fetchall()]
        logger.info(f"数据库表列: {existing_columns}")
        
        # 检查所需列是否存在，如不存在则添加
        required_columns = list(numeric_fields.values()) + ['dental_caries', 'physical_test_status']
        missing_columns = [col for col in required_columns if col not in existing_columns]
        
        if missing_columns:
            logger.warning(f"数据库表缺少列: {missing_columns}")
            for column in missing_columns:
                column_type = "TEXT" if column in ['dental_caries', 'physical_test_status'] else "REAL"
                logger.info(f"添加缺失的列: {column} ({column_type})")
                try:
                    cursor.execute(f"ALTER TABLE students ADD COLUMN {column} {column_type}")
                except sqlite3.Error as e:
                    logger.error(f"添加列 {column} 时出错: {str(e)}")
            
            conn.commit()
            logger.info("已添加缺失的列")
        
        # 构建更新SQL
        update_fields = []
        params = []
        
        # 基本字段
        update_fields.append("name = ?")
        params.append(data.get('name', ''))
        
        update_fields.append("gender = ?")
        params.append(data.get('gender', ''))
        
        update_fields.append("class = ?")
        params.append(data.get('class', ''))
        
        # 数值字段
        for db_field in numeric_fields.values():
            if db_field in existing_columns:
                update_fields.append(f"{db_field} = ?")
                params.append(processed_values.get(db_field))
        
        # 文本字段
        if 'dental_caries' in existing_columns:
            update_fields.append("dental_caries = ?")
            params.append(data.get('dental_caries', ''))
        
        if 'physical_test_status' in existing_columns:
            update_fields.append("physical_test_status = ?")
            params.append(data.get('physical_test_status', ''))
        
        # 更新时间
        update_fields.append("updated_at = ?")
        params.append(now)
        
        # 添加WHERE条件
        params.append(student_id)
        
        # 构建完整的SQL语句
        update_sql = f"UPDATE students SET {', '.join(update_fields)} WHERE id = ?"
        logger.info(f"更新SQL: {update_sql}")
        logger.info(f"参数: {params}")
        
        # 执行更新
        cursor.execute(update_sql, params)
        conn.commit()
        
        logger.info(f"成功更新学生信息: ID={student_id}, 名称={data.get('name', 'unknown')}")
        
        return jsonify({
            'status': 'ok',
            'message': f'成功更新学生信息: {data.get("name", student_id)}',
            'student_id': student_id
        })
    except sqlite3.Error as e:
        if conn:
            conn.rollback()
        error_msg = f"数据库错误: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        return jsonify({'error': error_msg}), 500
    except Exception as e:
        if conn:
            conn.rollback()
        error_msg = f"更新学生信息时出错: {str(e)}"
        logger.error(error_msg)
        logger.error(traceback.format_exc())
        return jsonify({'error': error_msg}), 500
    finally:
        if conn:
            conn.close()

# 删除学生API
@app.route('/api/students/<student_id>', methods=['DELETE'], strict_slashes=False)
def delete_student(student_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 检查学生是否存在
    cursor.execute('SELECT id, name FROM students WHERE id = ?', (student_id,))
    student = cursor.fetchone()
    
    if not student:
        conn.close()
        return jsonify({'error': f'未找到学号为 {student_id} 的学生'}), 404
    
    student_name = student['name']
    
    try:
        cursor.execute('DELETE FROM students WHERE id = ?', (student_id,))
        conn.commit()
        
        return jsonify({
            'status': 'ok',
            'message': f'成功删除学生: {student_name}',
            'student_id': student_id
        })
    except Exception as e:
        conn.rollback()
        return jsonify({'error': f'删除学生时出错: {str(e)}'}), 500
    finally:
        conn.close()

# 重置数据库API
@app.route('/api/reset-database', methods=['POST'])
def reset_database_api():
    # 安全检查，要求确认参数
    data = request.json or {}
    confirm = data.get('confirm', False)
    
    if not confirm:
        return jsonify({
            'status': 'error',
            'message': '请提供确认参数以重置数据库'
        }), 400
    
    success, message = reset_db()
    
    if success:
        return jsonify({
            'status': 'ok',
            'message': message
        })
    else:
        return jsonify({
            'status': 'error',
            'message': message
        }), 500

# 获取数据库信息API
@app.route('/api/database-info', methods=['GET'])
def database_info():
    try:
        # 获取数据库文件路径
        db_path = os.path.abspath(DATABASE)
        
        # 获取文件修改时间
        if os.path.exists(db_path):
            last_modified = datetime.datetime.fromtimestamp(os.path.getmtime(db_path)).strftime('%Y-%m-%d %H:%M:%S')
        else:
            last_modified = '数据库文件不存在'
        
        # 获取学生数量
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT COUNT(*) FROM students')
        student_count = cursor.fetchone()[0]
        conn.close()
        
        return jsonify({
            'status': 'ok',
            'path': db_path,
            'last_modified': last_modified,
            'student_count': student_count
        })
    except Exception as e:
        print(f"获取数据库信息时出错: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({
            'status': 'error',
            'message': f'获取数据库信息时出错: {str(e)}'
        }), 500

# 获取学生评语API
@app.route('/api/comments/<student_id>', methods=['GET'], strict_slashes=False)
def get_student_comment(student_id):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute('SELECT id, name, comments FROM students WHERE id = ?', (student_id,))
    student = cursor.fetchone()
    
    conn.close()
    
    if not student:
        return jsonify({
            'status': 'error',
            'message': '未找到该学生'
        }), 404
    
    return jsonify({
        'status': 'ok',
        'comment': {
            'studentId': student['id'],
            'studentName': student['name'],
            'content': student['comments'] or ''
        }
    })

# 保存学生评语API
@app.route('/api/comments', methods=['POST'], strict_slashes=False)
def save_student_comment():
    data = request.json
    
    if not data or 'studentId' not in data or 'content' not in data:
        return jsonify({
            'status': 'error',
            'message': '请提供学生ID和评语内容'
        }), 400
        
    student_id = data['studentId']
    content = data['content']
    append_mode = data.get('appendMode', False)
    
    logger.info(f"保存学生评语, 学生ID: {student_id}, 追加模式: {append_mode}")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # 先检查学生是否存在
    cursor.execute('SELECT id, comments FROM students WHERE id = ?', (student_id,))
    student = cursor.fetchone()
    
    if not student:
        conn.close()
        return jsonify({
            'status': 'error',
            'message': '未找到该学生'
        }), 404
    
    now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    try:
        # 如果是追加模式且已有评语，则在原有评语基础上添加新内容
        if append_mode and student['comments']:
            # 在原评语后添加新评语（添加时间戳和分隔符）
            updated_content = f"{student['comments']}\n\n--- {now} ---\n{content}"
        else:
            # 直接使用新评语
            updated_content = content
            
        # 更新学生评语
        cursor.execute('UPDATE students SET comments = ?, updated_at = ? WHERE id = ?', 
                      (updated_content, now, student_id))
        conn.commit()
        
        return jsonify({
            'status': 'ok',
            'message': '评语保存成功',
            'updatedContent': updated_content
        })
    except Exception as e:
        conn.rollback()
        logger.error(f"保存评语时出错: {str(e)}")
        return jsonify({
            'status': 'error',
            'message': f'保存评语时出错: {str(e)}'
        }), 500
    finally:
        conn.close()

# 获取评语模板API
@app.route('/api/comment-templates', methods=['GET'])
def get_comment_templates():
    """
    获取预定义的评语模板列表
    """
    templates = [
        # 学习类评语模板
        {"id": 1, "title": "品德优良", "content": "品德优良，尊敬师长，团结同学。", "type": "study"},
        {"id": 2, "title": "学习优秀", "content": "学习刻苦认真，上课认真听讲，作业按时完成。", "type": "study"},
        {"id": 3, "title": "学习积极", "content": "学习态度积极，能够主动思考，乐于探索新知识。", "type": "study"},
        {"id": 4, "title": "学习进步", "content": "近期学习有明显进步，在班级表现积极。", "type": "study"},
        {"id": 5, "title": "成绩优异", "content": "各科成绩优异，是班级的学习标兵。", "type": "study"},
        
        # 体育类评语模板
        {"id": 6, "title": "身体健康", "content": "身体健康，积极参加体育活动。", "type": "physical"},
        {"id": 7, "title": "运动技能", "content": "运动能力强，在体育活动中表现出色。", "type": "physical"},
        {"id": 8, "title": "体育精神", "content": "在体育活动中展现了团队合作精神和拼搏精神。", "type": "physical"},
        
        # 行为类评语模板
        {"id": 9, "title": "全面发展", "content": "德智体美劳全面发展，综合素质优秀。", "type": "behavior"},
        {"id": 10, "title": "行为规范", "content": "行为规范，能够严格遵守校规校纪。", "type": "behavior"},
        {"id": 11, "title": "积极参与", "content": "积极参与班级和学校活动，热心为集体服务。", "type": "behavior"},
        {"id": 12, "title": "有进步空间", "content": "在学习上有进步空间，希望能更加努力。", "type": "behavior"}
    ]
    
    return jsonify({
        "status": "ok",
        "templates": templates
    })

# 批量更新评语
@app.route('/api/batch-update-comments', methods=['POST'])
def batch_update_comments():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'status': 'error', 'message': '未接收到数据'})
        
        content = data.get('content', '').strip()
        append_mode = data.get('appendMode', True)
        student_ids = data.get('studentIds', [])  # 获取选中的学生ID列表
        
        if not content:
            return jsonify({'status': 'error', 'message': '评语内容不能为空'})
        
        if not student_ids:
            return jsonify({'status': 'error', 'message': '未选择学生'})
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # 记录更新的学生数量
        updated_count = 0
        
        # 获取当前时间，但不再添加到评语中
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        
        # 对选中的每个学生更新评语
        for student_id in student_ids:
            # 查询学生当前评语
            cursor.execute('SELECT comments FROM students WHERE id = ?', (student_id,))
            student = cursor.fetchone()
            
            if not student:
                continue  # 跳过不存在的学生ID
            
            current_comment = student[0] or ''
            
            # 根据模式设置新评语
            if append_mode and current_comment:
                # 简化追加模式，直接添加到末尾
                new_comment = f"{current_comment.strip()}\n\n{content}"
            else:
                # 如果是替换模式或无评语，则直接使用新内容
                new_comment = content
            
            # 更新学生评语和更新时间
            cursor.execute('UPDATE students SET comments = ?, updated_at = ? WHERE id = ?', 
                          (new_comment, now, student_id))
            
            updated_count += 1
        
        conn.commit()
        conn.close()
        
        return jsonify({
            'status': 'ok', 
            'message': f'已成功更新 {updated_count} 名学生的评语'
        })
        
    except Exception as e:
        logger.error(f"批量更新评语时出错: {str(e)}")
        return jsonify({'status': 'error', 'message': f'服务器错误: {str(e)}'})

# 导出评语为PDF
@app.route('/api/export-comments-pdf', methods=['GET'])
def api_export_comments_pdf():
    # 获取班级参数（可选）
    class_name = request.args.get('class')
    
    # 调用新的PDF导出函数
    result = export_comments_to_pdf(class_name)
    
    if result['status'] == 'ok':
        return jsonify(result)
    else:
        return jsonify(result), 500

# 提供导出文件下载
@app.route('/download/exports/<filename>', methods=['GET'])
def download_export(filename):
    return send_from_directory(EXPORTS_FOLDER, filename)

# 生成打印预览HTML
@app.route('/api/preview-comments', methods=['GET'])
def api_preview_comments():
    # 获取班级参数（可选）
    class_name = request.args.get('class')
    
    # 调用预览函数
    result = generate_preview_html(class_name)
    
    if result['status'] == 'ok':
        # 返回HTML内容而不是JSON
        return result['html']
    else:
        return jsonify(result), 500

# 学生成绩管理
# grades_manager已在文件开头初始化

# 获取所有学生成绩
@app.route('/api/grades', methods=['GET'])
def get_all_grades():
    try:
        semester = request.args.get('semester', '')
        if not semester:
            return jsonify({'status': 'error', 'message': '请提供学期'})
        
        grades = grades_manager.get_all_grades(semester)
        return jsonify({'status': 'ok', 'grades': grades})
    except Exception as e:
        app.logger.error(f'获取学生成绩时出错: {str(e)}')
        return jsonify({'status': 'error', 'message': f'获取学生成绩失败: {str(e)}'})

# 获取单个学生成绩
@app.route('/api/grades/<student_id>', methods=['GET'])
def get_student_grades(student_id):
    try:
        semester = request.args.get('semester', None)
        grades = grades_manager.get_student_grades(student_id, semester)
        return jsonify({'status': 'ok', 'grades': grades})
    except Exception as e:
        app.logger.error(f'获取学生成绩时出错: {str(e)}')
        return jsonify({'status': 'error', 'message': f'获取学生成绩失败: {str(e)}'})

# 保存学生成绩
@app.route('/api/grades/<student_id>', methods=['POST'])
def save_student_grade(student_id):
    try:
        data = request.get_json()
        app.logger.info(f"收到保存成绩请求，学生ID: {student_id}")
        app.logger.info(f"请求数据: {data}")
        
        semester = data.get('semester', '上学期')
        
        if not semester:
            app.logger.error("缺少必要参数: semester")
            return jsonify({'status': 'error', 'message': '缺少必要参数: semester'})
        
        # 移除semester，因为在save_grade函数中已经单独处理了
        grade_data = {}
        for field in data:
            if field != 'semester' and field in ['daof', 'yuwen', 'shuxue', 'yingyu', 'laodong', 'tiyu', 
                     'yinyue', 'meishu', 'kexue', 'zonghe', 'xinxi', 'shufa']:
                grade_data[field] = data.get(field, '')
        
        app.logger.info(f"提取的成绩数据: {grade_data}")
        success = grades_manager.save_grade(student_id, grade_data, semester)
        
        if success:
            app.logger.info(f"成功保存学生 {student_id} 的成绩")
            return jsonify({'status': 'ok', 'message': '成功保存学生成绩'})
        else:
            app.logger.error(f"保存学生 {student_id} 的成绩失败")
            return jsonify({'status': 'error', 'message': '保存学生成绩失败'})
    except Exception as e:
        app.logger.error(f'保存学生成绩时出错: {str(e)}')
        app.logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': f'保存学生成绩失败: {str(e)}'})

# 删除学生成绩
@app.route('/api/grades/<student_id>', methods=['DELETE'])
def delete_student_grade(student_id):
    try:
        semester = request.args.get('semester', '上学期')
        app.logger.info(f"收到删除成绩请求，学生ID: {student_id}, 学期: {semester}")
        
        success = grades_manager.delete_grade(student_id, semester)
        
        if success:
            app.logger.info(f"成功删除学生 {student_id} 的成绩")
            return jsonify({'status': 'ok', 'message': '成功删除学生成绩'})
        else:
            app.logger.error(f"删除学生 {student_id} 的成绩失败")
            return jsonify({'status': 'error', 'message': '删除学生成绩失败'})
    except Exception as e:
        app.logger.error(f'删除学生成绩时出错: {str(e)}')
        return jsonify({'status': 'error', 'message': f'删除学生成绩失败: {str(e)}'})

# 预览成绩导入
@app.route('/api/grades/preview-import', methods=['POST'])
def preview_grades_import():
    try:
        semester = request.form.get('semester', '上学期')
        app.logger.info(f"收到预览成绩导入请求，学期: {semester}")
        
        if 'file' not in request.files:
            app.logger.error("未提供文件")
            return jsonify({'status': 'error', 'message': '没有上传文件'}), 400
        
        file = request.files['file']
        app.logger.info(f"上传的文件名: {file.filename}")
        
        if file.filename == '':
            app.logger.error("未选择文件")
            return jsonify({'status': 'error', 'message': '未选择文件'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            app.logger.error("文件格式不正确")
            return jsonify({'status': 'error', 'message': '只能上传Excel (.xlsx/.xls) 文件'}), 400
        
        # 创建上传目录
        if not os.path.exists(UPLOAD_FOLDER):
            app.logger.info(f"创建上传目录: {UPLOAD_FOLDER}")
            os.makedirs(UPLOAD_FOLDER)
        
        # 保存文件
        timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
        saved_filename = f"{timestamp}_{secure_filename(file.filename)}"
        file_path = os.path.join(UPLOAD_FOLDER, saved_filename)
        file.save(file_path)
        app.logger.info(f"保存文件到: {file_path}")
        
        # 确认文件是否成功保存
        if not os.path.exists(file_path):
            app.logger.error(f"文件保存失败: {file_path}")
            return jsonify({'status': 'error', 'message': '文件保存失败'}), 500
        
        app.logger.info(f"文件大小: {os.path.getsize(file_path)} 字节")
        
        # 调用预览功能
        app.logger.info("开始预览成绩导入")
        result = grades_manager.preview_grades_from_excel(file_path, semester)
        
        if result['status'] == 'ok':
            app.logger.info(f"成功预览成绩: {result['message']}")
            return jsonify(result)
        else:
            app.logger.error(f"预览成绩失败: {result['message']}")
            return jsonify(result), 400
            
    except Exception as e:
        app.logger.error(f'预览成绩导入时出错: {str(e)}')
        app.logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error', 
            'message': f'预览成绩导入失败: {str(e)}'
        }), 500

# 确认导入成绩
@app.route('/api/grades/confirm-import', methods=['POST'])
def confirm_grades_import():
    try:
        data = request.json
        app.logger.info(f"收到确认导入成绩请求")
        
        if not data or 'file_path' not in data:
            app.logger.error("请求缺少文件路径")
            return jsonify({'status': 'error', 'message': '缺少文件路径参数'}), 400
        
        file_path = data['file_path']
        semester = data.get('semester', '上学期')
        
        # 导入成绩
        app.logger.info(f"开始导入成绩，文件路径: {file_path}, 学期: {semester}")
        success, message = grades_manager.import_grades_from_excel(file_path, semester)
        
        if success:
            app.logger.info(f"成功导入成绩: {message}")
            return jsonify({'status': 'ok', 'message': message})
        else:
            app.logger.error(f"导入成绩失败: {message}")
            return jsonify({'status': 'error', 'message': message}), 400
    
    except Exception as e:
        app.logger.error(f'确认导入成绩时出错: {str(e)}')
        app.logger.error(traceback.format_exc())
        return jsonify({
            'status': 'error', 
            'message': f'确认导入成绩失败: {str(e)}'
        }), 500

# 导入学生成绩
@app.route('/api/grades/import', methods=['POST'])
def import_grades():
    try:
        semester = request.form.get('semester', '上学期')
        app.logger.info(f"收到导入成绩请求，学期: {semester}")
        
        if 'file' not in request.files:
            app.logger.error("未提供文件")
            return jsonify({'status': 'error', 'message': '没有上传文件'})
        
        file = request.files['file']
        app.logger.info(f"上传的文件名: {file.filename}")
        
        if file.filename == '':
            app.logger.error("未选择文件")
            return jsonify({'status': 'error', 'message': '未选择文件'})
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            app.logger.error("文件格式不正确")
            return jsonify({'status': 'error', 'message': '只能上传Excel (.xlsx/.xls) 文件'})
        
        # 创建上传目录
        if not os.path.exists(UPLOAD_FOLDER):
            app.logger.info(f"创建上传目录: {UPLOAD_FOLDER}")
            os.makedirs(UPLOAD_FOLDER)
        
        # 保存文件
        saved_filename = secure_filename(file.filename)
        file_path = os.path.join(UPLOAD_FOLDER, saved_filename)
        file.save(file_path)
        app.logger.info(f"保存文件到: {file_path}")
        
        # 确认文件是否成功保存
        if not os.path.exists(file_path):
            app.logger.error(f"文件保存失败: {file_path}")
            return jsonify({'status': 'error', 'message': '文件保存失败'})
        
        app.logger.info(f"文件大小: {os.path.getsize(file_path)} 字节")
        
        # 调用预览功能而不是直接导入
        app.logger.info("改为使用预览功能")
        result = grades_manager.preview_grades_from_excel(file_path, semester)
        
        if result['status'] == 'ok':
            app.logger.info(f"成功预览成绩: {result['message']}")
            return jsonify(result)
        else:
            app.logger.error(f"预览成绩失败: {result['message']}")
            return jsonify(result), 400
            
    except Exception as e:
        app.logger.error(f'导入成绩时出错: {str(e)}')
        app.logger.error(traceback.format_exc())
        
        # 提供更明确的错误信息
        error_message = str(e)
        if "No such file or directory" in error_message:
            error_message = f"找不到文件或目录: {error_message}"
        elif "Permission denied" in error_message:
            error_message = f"权限被拒绝: {error_message}"
        
        return jsonify({'status': 'error', 'message': f'导入成绩失败: {error_message}'})

# 下载成绩导入模板
@app.route('/api/grades/template', methods=['GET'])
def download_grades_template():
    try:
        app.logger.info("收到下载成绩导入模板请求")
        template_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'templates', 'grades_import_template.xlsx')
        app.logger.info(f"查找模板文件: {template_path}")
        
        if not os.path.exists(template_path):
            app.logger.info("模板文件不存在，创建新模板")
            # 如果模板不存在，创建一个新的
            template_path = grades_manager.create_empty_template()
            app.logger.info(f"创建的新模板路径: {template_path}")
        
            if not template_path or not os.path.exists(template_path):
                app.logger.error("创建模板失败或模板路径无效")
                return jsonify({'status': 'error', 'message': '创建成绩导入模板失败'})
        
        app.logger.info(f"准备发送模板文件: {template_path}")
        return send_file(template_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='成绩导入模板.xlsx')
    except Exception as e:
        app.logger.error(f'下载成绩导入模板时出错: {str(e)}')
        app.logger.error(traceback.format_exc())
        return jsonify({'status': 'error', 'message': f'下载成绩导入模板失败: {str(e)}'})

# AI生成评语API
@app.route('/api/generate-comment', methods=['POST'])
def generate_comment():
    """生成AI评语"""
    try:
        # 获取请求数据
        data = request.get_json()
        print("收到评语生成请求:", data)
        
        # 验证请求数据
        comment_generator = CommentGenerator(deepseek_api.api_key)
        is_valid, error_msg = comment_generator.validate_request(data)
        if not is_valid:
            print("请求数据验证失败:", error_msg)
            return jsonify({
                "status": "error",
                "message": error_msg
            })
        
        # 获取学生信息
        student_id = data.get('student_id')
        if not student_id:
            print("缺少学生ID")
            return jsonify({
                "status": "error",
                "message": "缺少学生ID"
            })
            
        # 从数据库获取学生信息
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT name, gender FROM students WHERE id = ?', (student_id,))
        student = cursor.fetchone()
        conn.close()
        
        if not student:
            print(f"未找到学生: {student_id}")
            return jsonify({
                "status": "error",
                "message": "未找到学生"
            })
            
        # 构建学生信息字典
        student_info = {
            "name": student[0],
            "gender": student[1],
            "personality": data.get('personality', ''),
            "study_performance": data.get('study_performance', ''),
            "hobbies": data.get('hobbies', ''),
            "improvement": data.get('improvement', '')
        }
        
        # 格式化学生信息
        student_info = comment_generator.format_student_info(student_info)
        
        # 生成评语
        result = comment_generator.generate_comment(
            student_info=student_info,
            style=data.get('style', '鼓励性的'),
            tone=data.get('tone', '正式的'),
            max_length=int(data.get('max_length', 200))
        )
        
        print("评语生成结果:", result)
        
        # 返回结果
        if result["status"] == "ok":
            return jsonify({
                "status": "ok",
                "comment": result["comment"]
            })
        else:
            return jsonify({
                "status": "error",
                "message": result["message"]
            })
                
    except Exception as e:
        print("生成评语时出错:", str(e))
        print(traceback.format_exc())
        return jsonify({
            "status": "error",
            "message": f"生成评语时出错: {str(e)}"
        })

# 测试DeepSeek API连接
@app.route('/api/test-deepseek', methods=['POST'])
def test_deepseek_api():
    """测试DeepSeek API连接"""
    if not deepseek_api:
        return jsonify({
            "status": "error",
            "message": "DeepSeek API未初始化，请先设置API密钥"
        })
        
    try:
        result = deepseek_api.test_connection()
        return jsonify(result)
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"测试API连接时出错: {str(e)}"
        })

# 保存DeepSeek API设置
@app.route('/api/settings/deepseek', methods=['POST'])
def save_deepseek_api_settings():
    """保存DeepSeek API设置"""
    try:
        data = request.get_json()
        api_key = data.get('api_key', '').strip()
        
        # 更新全局设置
        global DEEPSEEK_API_KEY, deepseek_api, SYSTEM_SETTINGS
        DEEPSEEK_API_KEY = api_key
        SYSTEM_SETTINGS["deepseek_api_key"] = api_key
        
        # 重新初始化DeepSeek API
        if api_key:
            deepseek_api = DeepSeekAPI(api_key)
            SYSTEM_SETTINGS["deepseek_api_enabled"] = True
        else:
            deepseek_api = None
            SYSTEM_SETTINGS["deepseek_api_enabled"] = False
            
        return jsonify({
            "status": "ok",
            "message": "DeepSeek API设置已更新",
            "settings": SYSTEM_SETTINGS
        })
    except Exception as e:
        return jsonify({
            "status": "error",
            "message": f"保存设置时出错: {str(e)}"
        })

# 获取系统设置
@app.route('/api/settings', methods=['GET'])
def get_system_settings():
    """获取系统设置"""
    return jsonify({
        "status": "ok",
        "settings": SYSTEM_SETTINGS
    })

# 初始化数据应用
if __name__ == '__main__':
    # 初始化数据库，只确保表和列存在，不会重置数据
    init_db()
    logger.info("数据库初始化完成，保留已有数据")
    
    # 打印重要配置信息
    print("=============== 健康体检系统服务器 ===============")
    print(f"数据库路径: {os.path.abspath(DATABASE)}")
    print(f"上传文件夹: {os.path.abspath(UPLOAD_FOLDER)}")
    print(f"模板文件夹: {os.path.abspath(TEMPLATE_FOLDER)}")
    
    # 设置Flask应用程序
    is_production = os.environ.get('FLASK_ENV') == 'production'
    if is_production:
        # 生产环境设置
        app.run(host='0.0.0.0', port=8080)
    else:
        # 开发环境设置
        app.debug = True
        app.config['PROPAGATE_EXCEPTIONS'] = True
        app.run(host='0.0.0.0', port=8080) 

@app.route('/api/template')
def download_template():
    template_path = os.path.join(TEMPLATE_FOLDER, 'student_template.xlsx')
    if not os.path.exists(template_path):
        create_student_template()
    
    return jsonify({
        'status': 'ok',
        'url': url_for('download_student_template')
    })